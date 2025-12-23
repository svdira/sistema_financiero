import os
import sqlite3
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook, Workbook

def fila_concepto(ws):
    for r in range(1,26):
        val = ws.cell(row=r, column=1).value
        if isinstance(val,str) and val.strip().lower() in ["concepto","indicador"]:
            return r
    return None

def fila_es_vacia(ws,fila):
    max_col = ws.max_column or 1
    for c in range(1, max_col + 1):
        v = ws.cell(row=fila, column = c).value
        if v is None:
            continue
        if isinstance(v,str):
            if v.strip() != "":
                return False
        else:
            return False
    return True

def procesar_xlsx(path_src, path_dst_dir):
    wb = load_workbook(filename=path_src, data_only=False)
    for ws in wb.worksheets:
        r = fila_concepto(ws)
        if r is not None and r > 1:
            ws.delete_rows(idx=1, amount = r - 1)

        limite = min(350, ws.max_row or 0)
        for i in range(limite,0,-1):
            if fila_es_vacia(ws,i):
                ws.delete_rows(idx=i, amount=1)
 
    ruta_dst = Path(path_dst_dir) / Path(path_src).name
    wb.save(ruta_dst)

def es_vacio(val):
    if val is None:
        return True
    if isinstance(val,str) and val.strip()=="":
        return True
    return False

def ultima_fila(ws):
    max_row = ws.max_row or 1
    for r in range(max_row,1,-1):
        v = ws.cell(row=r,column=1).value
        if not es_vacio(v):
            return r
    return 2

def ultima_columna(ws,fila):
    max_col = ws.max_column or 1
    for c in range(max_col, 0, -1):
        v = ws.cell(row=fila, column=c).value
        if not es_vacio(v):
            return c
    return 1

def unmerge_all(ws):
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))

def find_no_in_A1_A25(ws):
    for r in range(1,26):
        if ws.cell(row=r,column=1).value == "No.":
            return r
    return None

def is_empty_row(ws, row_idx: int) -> bool:
    for cell in ws[row_idx]:
        v = cell.value
        if v in None:
            continue
        if isinstance(v,str) and v.strip() == "":
            continue
        return False
    return True

def process_xlsx(src_path: Path, dest_dir: Path):
    wb = load_workbook(src_path, data_only = True)
    for ws in wb.worksheets:
        unmerge_all(ws)
        found_row = find_no_in_A1_A25(ws)
        if found_row and found_row > 1:
            ws.delete_rows(1,found_row -1)
        for r in range(min(ws.max_row,350), 0, -1):
            if is_empty_row(ws, r):
                ws.delete_rows(r)
    wb.save(dest_dir / src_path.name)

def process_xls(src_path: Path, dest_dir: Path):
    xls = pd.ExcelFile(src_path)
    out_path = dest_dir / (src_path.stem + ".xlsx")
    with pd.ExcelWriter(out_path, engine = "openpyxl") as writer:
        for sheet_name in xls.sheet_names:
            df = xls.parse(sheet_name = sheet_name, header = None)
            found_idx = None
            for r in range(min(25, len(df))):
                if str(df.iat[r,0]).strip() == "No.":
                    found_idx = r
                    break
            if found_idx and found_idx > 0:
                df = df.iloc[found_idx:]
            head_end = min(350,len(df))
            head = df.iloc[:head_end]
            mask_empty = head.apply(lambda row: all(str(x).strip() == "" or pd.isna(x) for x in row), axis=1)
            head_clean = head.loc[~mask_empty]
            tail = df.iloc[head_end:]
            df_clean = pd.concat([head_clean, tail], ignore_index = True)
            df_clean.to_excel(writer, sheet_name=sheet_name, header = False, index=False)
    
def procesar_informe(path_src):
    wb = load_workbook(filename=path_src, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    ultima_col = ultima_columna(ws,fila=1)
    ultima_row = ultima_fila(ws)

    registros = []

    for col in range(2,max(2,ultima_col)+1):
        banco = ws.cell(row=1,column=col).value
        for fila in range(2, max(2, ultima_row) + 1):
            concepto = ws.cell(row=fila, column=1).value
            valor = ws.cell(row=fila,column=col).value
            if es_vacio(concepto) and es_vacio(valor):
                continue
            registros.append({
                "Archivo":Path(path_src).name,
                "Banco": banco,
                "Concepto": concepto,
                "Valor":valor
            })
    wb.close()

    return registros

def last_col_in_row(ws, row, start_col=1):
    col = start_col
    while ws.cell(row=row, column=col).value not in (None, ""):
        col += 1
    return col - 1

def last_row_in_col(ws, col, start_row=1):
    row = start_row
    while ws.cell(row=row, column=col).value not in (None, ""):
        row += 1
    return row - 1
    
def consolidar_carteras(destino: Path, archivos):
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S%f")[:-3]
    out_path = destino / f"cartera_clasificada_{timestamp}.xlsx"
    wb_con = Workbook()
    ws_con = wb_con.active
    ws_con.title = "consolidado"

    headers = ["Archivo","Institucion","Categoria","Monto","Saldo","NoCreditos","Rubro"]
    ws_con.append(headers)

    ultima_disp = 2

    for file in archivos:
        file_name = file.split("/")[-1]
        ruta_archivo = f"raw_ssf/cartera_formateados/{file}"
        wb = load_workbook(ruta_archivo, data_only = True)
        for ws in wb.worksheets:
            ultima_col = last_col_in_row(ws, row=2, start_col=3)
            ultima_fila = last_row_in_col(ws, col=2, start_row=3)

            for intC in range(3,ultima_col - 1, 3):
                categorias = [ws.cell(row=r,column=2).value for r in range(3,ultima_fila +1)]
                bloque = []
                for r in range(3,ultima_fila + 1):
                    bloque.append([
                        ws.cell(row=r, column=intC).value,
                        ws.cell(row=r, column=intC+1).value,
                        ws.cell(row=r, column=intC+2).value
                    ])
                institucion = ws.cell(row=1, column=intC).value

                this_r = "No Class"
                for idx, cat in enumerate(categorias):   
                    monto, saldo, nocred = bloque[idx]
                    this_r = cat if len(str(cat))>2 else this_r
                    rubro = cat if cat and len(str(cat))>2 else this_r
                    ws_con.append([file_name,institucion,cat, monto,saldo,nocred,rubro])
        wb.close()
    wb_con.save(out_path)
    return out_path


def consolidar_estados(destino: Path, archivos, tipo):
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S%f")[:-3]
    out_path = destino / f"{tipo}_{timestamp}.xlsx"

    registros_totales = []
    extensiones = {".xlsx",".xlsm"}
    for arc in archivos:
        try:
            ruta_arc = f"raw_ssf/{tipo}_formateados/{arc}"
            registros = procesar_informe(ruta_arc)
            registros_totales.extend(registros)
        except Exception as e:
            print(f"[ERROR] {arc}: {e}")
    if not registros_totales:
        print("No se generaron registros. Validar.")
        return
    df = pd.DataFrame(registros_totales, columns = ["Archivo","Banco","Concepto","Valor"])
    with pd.ExcelWriter(out_path, engine="openpyxl") as Writer:
        df.to_excel(Writer, sheet_name="consolidado", index=False)

    print("Se terminó la consolidación.")
    return out_path

def listar_excels(carpeta: str):
    base = Path(carpeta)
    return [
        p.name
        for p in base.iterdir()
        if p.suffix.lower() in ('.xlsx', '.xls')
        and not p.name.startswith('~$')
    ]


def formatear_reportes(reporte):
    reportes_raw = listar_excels(f'raw_ssf/{reporte}')
    reportes_formateados = listar_excels(f'raw_ssf/{reporte}_formateados')
    faltantes = list(set(reportes_raw) - set(reportes_formateados))
    
    if len(faltantes) > 0:
        for f in faltantes:
            procesar_xlsx(f"raw_ssf/{reporte}/{f}",f'raw_ssf/{reporte}_formateados')
        print("Se formatearon los siguientes archivos: " +', '.join(faltantes))
    else:
        print("No hay archivos nuevos para procesar.")


def formatear_reportes_catera():
    reportes_raw = listar_excels(f'raw_ssf/cartera')
    raw_xlsx = []
    for a in reportes_raw:
        p = Path(a)
        if p.suffix == '.xls':
            raw_xlsx.append(p.with_suffix('.xlsx').name)
        else:
            raw_xlsx.append(a)
 
            
    reportes_formateados = listar_excels(f'raw_ssf/cartera_formateados')
    faltantes = list(set(raw_xlsx) - set(reportes_formateados))
    
    if len(faltantes) > 0:
        for f in faltantes:
            extension = Path(f"raw_ssf/cartera/{f}").suffix
            if extension == ".xlsx":
                process_xlsx(f"raw_ssf/cartera/{f}",f'raw_ssf/cartera_formateados')
            elif extension == ".xls":
                process_xls(Path(f"raw_ssf/cartera/{f}"),Path(f'raw_ssf/cartera_formateados'))
            else:
                return   
                
        print("Se formatearon los siguientes archivos: " +', '.join(faltantes))
    else:
        print("No hay archivos nuevos para procesar.")

def consolidar_publicar(reporte):
    con = sqlite3.connect('raw_ssf/sistema_financiero.db')
    cur = con.cursor()
    cur.execute(f"SELECT distinct Archivo FROM reportes WHERE tipo_reporte='{reporte}' ")
    filas = cur.fetchall()
    con.close()

    reportes_raw = listar_excels(f'raw_ssf/{reporte}_formateados')
    reportes_formateados = [f[0] for f in filas]
    faltantes = list(set(reportes_raw) - set(reportes_formateados))

    ruta = consolidar_estados(Path(f"raw_ssf/{reporte}_consolidados"),faltantes,reporte)
    print(ruta)
    if ruta:
        log_query = "INSERT INTO cargas_ssf (descripcion) VALUES (?) "
        with sqlite3.connect('raw_ssf/sistema_financiero.db') as con:
            cur = con.cursor()
            cur.execute(log_query,(reporte,))
            con.commit()
            ultimo_id = cur.lastrowid
        
        df = pd.read_excel(ruta)
        df.insert(0,'tipo_reporte', reporte)
        df.insert(0,'log_id', ultimo_id)
        with sqlite3.connect('raw_ssf/sistema_financiero.db') as con:
            df.to_sql('reportes', con, if_exists='append', index=False)
    
    else:
        print("No hay datos nuevos para insertar")

def consolidar_publicar_cartera():
    reporte = 'cartera'
    con = sqlite3.connect('raw_ssf/sistema_financiero.db')
    cur = con.cursor()
    cur.execute(f"SELECT distinct Archivo FROM carteras")
    filas = cur.fetchall()
    con.close()

    reportes_raw = listar_excels(f'raw_ssf/{reporte}_formateados')
    reportes_formateados = [f[0] for f in filas]
    faltantes = list(set(reportes_raw) - set(reportes_formateados))

    ruta = consolidar_carteras(Path(f"raw_ssf/{reporte}_consolidados"),faltantes)

    if ruta:
        log_query = "INSERT INTO cargas_ssf (descripcion) VALUES (?) "
        with sqlite3.connect('raw_ssf/sistema_financiero.db') as con:
            cur = con.cursor()
            cur.execute(log_query,(reporte,))
            con.commit()
            ultimo_id = cur.lastrowid
        
        df = pd.read_excel(ruta)
        df.insert(0,'log_id', ultimo_id)
        with sqlite3.connect('raw_ssf/sistema_financiero.db') as con:
            df.to_sql('carteras', con, if_exists='append', index=False)
    
    else:
        print("No hay datos nuevos para insertar")


formatear_reportes('balances')
formatear_reportes('indicadores')
formatear_reportes('resultados')
formatear_reportes_catera()


consolidar_publicar('balances')
consolidar_publicar('resultados')
consolidar_publicar('indicadores')
consolidar_publicar_cartera()